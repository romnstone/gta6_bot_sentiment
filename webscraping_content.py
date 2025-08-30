import requests
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

file_path = "extra_reddit_comments_GTA6.xlsx"
sheet_name = "Sheet1"

# urlList = ['https://www.reddit.com/r/GTA6/comments/1jvz51e/im_tired_of_waiting_for_trailer_2_or_more/.json', 'https://www.reddit.com/r/GTA6/comments/18afb01/everyone_is_copying_the_gta6_trailer_announcement/.json', 'https://www.reddit.com/r/SubredditDrama/comments/1hnnq8h/im_the_only_one_with_a_brain_here_rgta6_goes_into/.json',
#            'https://www.reddit.com/r/NCAAFBseries/comments/18azjtg/gta6_drops_a_trailer_for_a_game_2_years_away_and/.json', 'https://www.reddit.com/r/GTA6/comments/1890uf9/are_we_ready_for_the_70_things_you_missed_in_the/.json', 'https://www.reddit.com/r/TESVI/comments/1kgmvcb/now_that_we_seen_gta6_update_trailer/.json',
#            'https://www.reddit.com/r/GTA6/comments/1hilmv3/were_closer_to_the_release_of_gta6_than_we_are_to/.json', 'https://www.reddit.com/r/nederlands/comments/18b75ct/stop_dit_is_belangrijk_de_trailer_van_gta6/.json', 'https://www.reddit.com/r/lanoire/comments/1khk1ko/a_la_noire_sequel_is_looking_like_a_really_good/.json',
#            'https://www.reddit.com/r/GTA6/comments/1hsu6ur/guess_the_trailer_2_release_date_contest/.json', 'https://www.reddit.com/r/GTA6/comments/1jtrwuo/hypothetical_gta6_will_release_april_30th_with_no/.json', 'https://www.reddit.com/r/gaming/comments/18baoui/why_is_everyone_losing_their_shit_about_gta6/.json',
#            'https://www.reddit.com/r/whennews/comments/1kg4qw8/gta6_trailer_2_just_dropped/.json', 'https://www.reddit.com/r/PakGamers/comments/1kg6ync/your_thoughts_of_gta6_trailer_2/.json', 'https://www.reddit.com/r/GTA6/comments/124mnqn/plans_for_when_gta6_trailer_releases/.json',
#            'https://www.reddit.com/r/HalfLife/comments/1kg6nn3/we_got_gta6_trailer_2_before_hl3/.json', 'https://www.reddit.com/r/rockstar/comments/1kjuykq/if_rockstar_would_release_a_60fps_patch_for_rdr2/.json', 'https://www.reddit.com/r/NintendoSwitch2/comments/1jiy5zy/if_gta6_trailer_2_gets_released_around_the_switch/.json',
#            'https://www.reddit.com/r/GTA6/comments/fzgbnz/one_day_were_gonna_wake_up_and_the_real_gta6/.json', 'https://www.reddit.com/r/loveafterporn/comments/1kg5xe0/new_gta6_trailer/.json', 'https://www.reddit.com/r/vargskelethor/comments/1ki41p2/guys_does_anyone_know_what_song_was_playing/.json',
#            'https://www.reddit.com/r/GTA6_NEW/comments/1hf9ee2/if_gta6_trailer_2_does_not_release_on_december_17/.json', 'https://www.reddit.com/r/ps6/comments/1kj3zam/ps6_leaked_in_gta6_trailer_2/.json', 'https://www.reddit.com/r/CBSE/comments/1kgbdgr/we_got_gta6_trailer_2_before_cbse_2025_result/.json',
#            'https://www.reddit.com/r/80smusic/comments/1kgln4i/wang_chungs_everybody_have_fun_tonight_and_hot/.json', 'https://www.reddit.com/r/gtaonline/comments/188d11z/gta_6_trailer_release_date_december_5th/.json', 'https://www.reddit.com/r/GrandTheftAutoV/comments/1kgdbn0/the_gta6_2_trailers_soundtrack_pure_cinematic_gold/.json', 
#            'https://www.reddit.com/r/miniminter/comments/1ki1bw5/gta6_trailer_2/.json', 'https://www.reddit.com/r/CBSE/comments/1kg57dj/we_got_gta6_second_trailer_before_our_result/.json', 'https://www.reddit.com/r/GTA/comments/1khkurn/itt_speculate_on_what_gta6s_flaws_might_be/.json', 
#            'https://www.reddit.com/r/GTA/comments/1kgda2c/the_gta6_2_trailers_soundtrack_pure_cinematic_gold/.json', 'https://www.reddit.com/r/GTA/comments/1ki192d/cool_gta6_detail_from_trailer/.json', 'https://www.reddit.com/r/GTA6/comments/1ja66cf/my_favorite_scene_in_the_gta6_trailer/.json', 
#            'https://www.reddit.com/r/GTA/comments/1kgct3i/the_new_trailer_just_made_me_realize_something/.json', 'https://www.reddit.com/r/GTA6/comments/1k7b205/the_gta_6_trailer_was_playing_on_a_gas_pump/.json', 'https://www.reddit.com/r/GTA6/comments/1k6igcg/gta_6_trailer_1_is_featured_in_youtubes_20_year/.json', 
#            'https://www.reddit.com/r/MichaelDoesLife/comments/1hdeqlq/dont_need_or_have_to_watch_his_tga_stream_but_ill/.json', 'https://www.reddit.com/r/GTA6_NEW/comments/1h5b7pf/theres_no_announcement_for_gta6_trailer_2_and/.json', 'https://www.reddit.com/r/gtavcustoms/comments/18ay3j0/gta6_trailer_just_dropped_and_there_are_a_lot_of/.json', 
#            'https://www.reddit.com/r/GTAV/comments/1j5uua2/when_all_the_world_await_a_new_trailer_for/.json', 'https://www.reddit.com/r/playstation/comments/1gghh3h/after_the_gta6_trailer_i_wanted_to_buy_a_ps5_for/.json', 'https://www.reddit.com/r/GTA6unmoderated/comments/1j4ygjf/gta6_trailer_2_may_drop_april_8th/.json']

urlList = ['https://www.reddit.com/r/GTA6/comments/1jvz51e/im_tired_of_waiting_for_trailer_2_or_more/.json', 'https://www.reddit.com/r/GTA6/comments/18afb01/everyone_is_copying_the_gta6_trailer_announcement/.json', 'https://www.reddit.com/r/SubredditDrama/comments/1hnnq8h/im_the_only_one_with_a_brain_here_rgta6_goes_into/.json', 'https://www.reddit.com/r/NCAAFBseries/comments/18azjtg/gta6_drops_a_trailer_for_a_game_2_years_away_and/.json', 'https://www.reddit.com/r/GTA6/comments/1890uf9/are_we_ready_for_the_70_things_you_missed_in_the/.json', 'https://www.reddit.com/r/TESVI/comments/1kgmvcb/now_that_we_seen_gta6_update_trailer/.json', 'https://www.reddit.com/r/GTA6/comments/1hilmv3/were_closer_to_the_release_of_gta6_than_we_are_to/.json', 'https://www.reddit.com/r/nederlands/comments/18b75ct/stop_dit_is_belangrijk_de_trailer_van_gta6/.json', 'https://www.reddit.com/r/lanoire/comments/1khk1ko/a_la_noire_sequel_is_looking_like_a_really_good/.json', 'https://www.reddit.com/r/GTA6/comments/1hsu6ur/guess_the_trailer_2_release_date_contest/.json', 'https://www.reddit.com/r/GTA6/comments/1jtrwuo/hypothetical_gta6_will_release_april_30th_with_no/.json', 'https://www.reddit.com/r/gaming/comments/18baoui/why_is_everyone_losing_their_shit_about_gta6/.json', 'https://www.reddit.com/r/rockstar/comments/1kjuykq/if_rockstar_would_release_a_60fps_patch_for_rdr2/.json', 'https://www.reddit.com/r/whennews/comments/1kg4qw8/gta6_trailer_2_just_dropped/.json', 'https://www.reddit.com/r/GTA6/comments/124mnqn/plans_for_when_gta6_trailer_releases/.json', 'https://www.reddit.com/r/PakGamers/comments/1kg6ync/your_thoughts_of_gta6_trailer_2/.json', 'https://www.reddit.com/r/NintendoSwitch2/comments/1jiy5zy/if_gta6_trailer_2_gets_released_around_the_switch/.json', 'https://www.reddit.com/r/GTA6/comments/fzgbnz/one_day_were_gonna_wake_up_and_the_real_gta6/.json', 'https://www.reddit.com/r/GTA6_NEW/comments/1hf9ee2/if_gta6_trailer_2_does_not_release_on_december_17/.json', 'https://www.reddit.com/r/ps6/comments/1kj3zam/ps6_leaked_in_gta6_trailer_2/.json', 'https://www.reddit.com/r/vargskelethor/comments/1ki41p2/guys_does_anyone_know_what_song_was_playing/.json', 'https://www.reddit.com/r/gtaonline/comments/188d11z/gta_6_trailer_release_date_december_5th/.json', 'https://www.reddit.com/r/CBSE/comments/1kgbdgr/we_got_gta6_trailer_2_before_cbse_2025_result/.json', 'https://www.reddit.com/r/80smusic/comments/1kgln4i/wang_chungs_everybody_have_fun_tonight_and_hot/.json', 'https://www.reddit.com/r/GrandTheftAutoV/comments/1kgdbn0/the_gta6_2_trailers_soundtrack_pure_cinematic_gold/.json', 'https://www.reddit.com/r/miniminter/comments/1ki1bw5/gta6_trailer_2/.json', 'https://www.reddit.com/r/CBSE/comments/1kg57dj/we_got_gta6_second_trailer_before_our_result/.json', 'https://www.reddit.com/r/GTA/comments/1khkurn/itt_speculate_on_what_gta6s_flaws_might_be/.json', 'https://www.reddit.com/r/GTA/comments/1kgda2c/the_gta6_2_trailers_soundtrack_pure_cinematic_gold/.json', 'https://www.reddit.com/r/GTA6/comments/1ja66cf/my_favorite_scene_in_the_gta6_trailer/.json', 'https://www.reddit.com/r/GTA/comments/1ki192d/cool_gta6_detail_from_trailer/.json', 'https://www.reddit.com/r/GTA/comments/1kgct3i/the_new_trailer_just_made_me_realize_something/.json', 'https://www.reddit.com/r/MichaelDoesLife/comments/1hdeqlq/dont_need_or_have_to_watch_his_tga_stream_but_ill/.json', 'https://www.reddit.com/r/gtavcustoms/comments/18ay3j0/gta6_trailer_just_dropped_and_there_are_a_lot_of/.json', 'https://www.reddit.com/r/GTA6/comments/1k7b205/the_gta_6_trailer_was_playing_on_a_gas_pump/.json', 'https://www.reddit.com/r/GTA6/comments/1k6igcg/gta_6_trailer_1_is_featured_in_youtubes_20_year/.json', 'https://www.reddit.com/r/GTA6_NEW/comments/1h5b7pf/theres_no_announcement_for_gta6_trailer_2_and/.json', 'https://www.reddit.com/r/GTAV/comments/1j5uua2/when_all_the_world_await_a_new_trailer_for/.json', 'https://www.reddit.com/r/playstation/comments/1gghh3h/after_the_gta6_trailer_i_wanted_to_buy_a_ps5_for/.json', 'https://www.reddit.com/r/GTA6unmoderated/comments/1j4ygjf/gta6_trailer_2_may_drop_april_8th/.json', 'https://www.reddit.com/r/GTA6/comments/1j509jw/fans_dont_have_respect_for_the_ones_that_died/.json', 'https://www.reddit.com/r/explainlikeimfive/comments/1kglllf/eli5_why_do_videogame_environments_have_an_easier/.json', 'https://www.reddit.com/r/GamingLeaksAndRumours/comments/18aok8d/gta6_trailer_leak/.json', 'https://www.reddit.com/r/GTA6/comments/107y0pc/twitter_user_who_revealed_lucias_name_before_the/.json', 'https://www.reddit.com/r/GTA6/comments/1ae08r2/i_believe_the_2nd_trailer_for_gta6_will_come_out/.json', 'https://www.reddit.com/r/GTA6/comments/17gznyz/gta6_teasertrailer_on_feb_4_2024/.json', 'https://www.reddit.com/r/PS5/comments/1kghmrw/grand_theft_auto_vi_everything_you_need_to_know/.json', 'https://www.reddit.com/r/gaming/comments/188m2ft/rockstar_revealed_an_announcement_for_a_trailer/.json', 'https://www.reddit.com/r/playstation/comments/1io9cqg/valantine_day_gta6_trailer/.json', 'https://www.reddit.com/r/GTA6/comments/1ij0bu4/grand_theft_auto_vi_trailer_1_if_made_in_1989/.json', 'https://www.reddit.com/r/GamingLeaksAndRumours/comments/1h4m9wg/rockstar_games_updates_their_gta_vi_youtube/.json', 'https://www.reddit.com/r/thebottlemen/comments/1hfeqhq/which_will_come_first_catb_update_or_gta6_trailer/.json', 'https://www.reddit.com/r/kindafunny/comments/1i331c1/the_first_red_dead_redemption_2_trailer_came_not/.json', 'https://www.reddit.com/r/GTA/comments/1h5b6y3/theres_no_announcement_for_gta6_trailer_2_and/.json', 'https://www.reddit.com/r/CyberBoiUK/comments/1hnu1t2/little_gta6_trailer_2_poem/.json', 'https://www.reddit.com/r/SubredditDrama/comments/18p3d5p/the_gta6_hacker_is_institutionalized_indefinitely/.json', 'https://www.reddit.com/r/gamesEcultura/comments/188d3w2/quantos_views_o_trailer_de_gta6_vai_pegar_na/.json', 'https://www.reddit.com/r/italygames/comments/1efta46/cosa_ne_pensate_del_trailer_ufficiale_di_gta6/.json', 'https://www.reddit.com/r/GTA6/comments/18chgr9/gta6_stole_the_hype_from_the_finals_with_just_a/.json', 'https://www.reddit.com/r/DragonBallXenoverse2/comments/1clgjr9/fck_sparking_zero_when_xeno3_trailer_pulls_up_its/.json', 'https://www.reddit.com/r/GTA6/comments/18ay7gc/devs_sharing_their_thoughts_sad_that_this_happened/.json', 'https://www.reddit.com/r/GTA6/comments/xxgq5b/gta6_first_trailer_predictions/.json', 'https://www.reddit.com/r/GTA/comments/1et1dyo/what_is_one_word_you_can_describe_the_gta6/.json', 'https://www.reddit.com/r/PS5pro/comments/1kgfzjq/how_is_it_even_running/.json', 'https://www.reddit.com/r/GTA6/comments/1cngrhq/why_would_the_time_gaps_between_gta6_trailers_be/.json', 'https://www.reddit.com/r/GTA6/comments/17grbv6/official_26th_october_2023_thread/.json', 'https://www.reddit.com/r/GTA/comments/1eqcmc7/r_will_probably_change_some_things_that_were_in/.json', 'https://www.reddit.com/r/GTA6/comments/18xnbic/why_is_there_so_many_planes_in_gta6_trailer/.json']
usernames = []

# Reddit post in JSON format
# url = "https://www.reddit.com/r/climatechange/comments/1brgcx0/are_climate_change_model_error_margins_too_great/.json"
headers = {'User-agent': 'Mozilla/5.0'}

for url in urlList:
    response = requests.get(url, headers=headers)
    data = response.json()

    # The comments are in the second element of the JSON array
    thread_title = data[0]['data']['children'][0]['data']['title']

    comments = data[1]['data']['children']

    rows = []

    for comment in comments:
        if comment['kind'] != 't1':
            continue  # Skip if not a comment

        comment_data = comment['data']
        user = comment_data.get('author')
        timestamp = datetime.fromtimestamp(comment_data.get('created_utc'))
        content = comment_data.get('body')

        if user != '[deleted]':
            # Add to list of rows
            rows.append({
                'User': user,
                'Timestamp': timestamp,
                'Main Thread': thread_title,
                'Content': content
            })

            usernames.append(user)

    # Create dataframe
    df = pd.DataFrame(rows)

    # RUN THIS CODE: to create an file for the reddit comments
    # df.to_excel("reddit_comments_GTA6.xlsx", index=False)

    # RUN THIS CODE: to append to an existing file for the reddit comments
    # Load the workbook and get the last row
    book = load_workbook(file_path)
    sheet = book[sheet_name]
    last_row = sheet.max_row

    # Append to the sheet using startrow
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=last_row)

    print("Data appended to the last row of the existing sheet.")

    # print(usernames)